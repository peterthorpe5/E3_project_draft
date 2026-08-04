"""Compact integration proof for downloaded evidence through app-ready release."""

from __future__ import annotations

import hashlib
import json
import shutil
import sqlite3
from dataclasses import replace
from pathlib import Path
from typing import Any, Sequence

import duckdb
from openpyxl import load_workbook

from e3workflow.config import load_config
from e3workflow.integration import _final_query, run_app_ready_stage, run_integrated_stage
from e3workflow.ligandability import (
    POCKET_CONSERVATION_COLUMN_TYPES,
    POCKET_CONSERVATION_FIELDS,
    run_ligandability_stage,
)
from e3workflow.prioritisation import run_prestructure_stage
from e3workflow.production import run_domain_stage, run_expression_stage
from e3workflow.resources import (
    build_domain_cache_manifest,
    build_expression_manifest,
    build_ligandability_manifest,
)
from e3workflow.tabular import write_records


def _write_parquet(
    path: Path, schema: str, rows: Sequence[Sequence[Any]]
) -> None:
    """Write a typed Parquet fixture."""
    path.parent.mkdir(parents=True, exist_ok=True)
    connection = duckdb.connect(":memory:")
    try:
        connection.execute(f"CREATE TABLE fixture ({schema})")
        if rows:
            placeholders = ", ".join("?" for _ in rows[0])
            connection.executemany(
                f"INSERT INTO fixture VALUES ({placeholders})", rows
            )
        escaped = str(path).replace("'", "''")
        connection.execute(f"COPY fixture TO '{escaped}' (FORMAT PARQUET)")
    finally:
        connection.close()


def _membership_row(
    species: str, raw_identifier: str, accession: str, entry: str
) -> tuple[Any, ...]:
    """Return one hierarchical-membership row."""
    return (
        "HIERARCHICAL_ORTHOGROUP",
        "N0.HOG0001",
        "OG0001",
        "",
        species,
        raw_identifier,
        accession,
        entry,
        "reviewed",
        "fixture",
        "PARSED",
        "fixture",
        "fixture.tsv",
        2,
    )


def _annotation(accession: str) -> dict[str, Any]:
    """Return one terminal InterPro cache payload with F-box support."""
    return {
        "schema_version": 1,
        "requested_accession": accession,
        "retrieval_status": "ANNOTATED",
        "retrieved_at_utc": "2026-07-22T00:00:00Z",
        "api_base_url": "https://www.ebi.ac.uk/interpro/api",
        "release": {"interpro_version": "109.0"},
        "protein_metadata": None,
        "results": [
            {
                "metadata": {
                    "accession": "PF00646",
                    "name": "F-box domain",
                    "source_database": "pfam",
                    "type": "domain",
                    "integrated": "IPR001810",
                },
                "proteins": [
                    {
                        "accession": accession.lower(),
                        "protein_length": 8,
                        "source_database": "reviewed",
                        "organism": "fixture",
                        "in_alphafold": True,
                        "entry_protein_locations": [
                            {
                                "fragments": [
                                    {
                                        "start": 2,
                                        "end": 6,
                                        "dc-status": "CONTINUOUS",
                                    }
                                ],
                                "representative": False,
                                "model": "PF00646",
                                "score": 1e-8,
                            }
                        ],
                    }
                ],
            }
        ],
        "error": "",
    }


def test_final_query_accepts_legacy_empty_varchar_conservation_table(
    synthetic_config: Path,
    tmp_path: Path,
) -> None:
    """Stage 10 must accept empty Stage 09 Parquet files created before typed schemas."""
    prestructure = tmp_path / "prestructure.parquet"
    _write_parquet(
        prestructure,
        (
            "cluster_id VARCHAR, primary_group_type VARCHAR, primary_group_id VARCHAR, "
            "orthofinder_orthogroup_ids VARCHAR, "
            "orthofinder_hierarchical_group_ids VARCHAR, candidate_accessions VARCHAR, "
            "prestructure_score DOUBLE, target_species_total BIGINT, "
            "target_species_fraction DOUBLE, mandatory_species_fraction DOUBLE, "
            "domain_species_fraction DOUBLE, expression_species_fraction DOUBLE, "
            "grant_aligned_stringent_pass BOOLEAN, evidence_completeness_fraction DOUBLE, "
            "inclusion_reasons VARCHAR, exclusion_reasons VARCHAR, missing_evidence VARCHAR, "
            "profile_name VARCHAR"
        ),
        [
            (
                "cluster_legacy_empty",
                "HIERARCHICAL_ORTHOGROUP",
                "N0.HOG0001",
                "OG0001",
                "N0.HOG0001",
                "Q9SA03",
                0.8,
                2,
                1.0,
                1.0,
                1.0,
                1.0,
                True,
                1.0,
                "fixture",
                "",
                "structural evidence unavailable",
                "fixture_profile",
            )
        ],
    )
    conservation = tmp_path / "legacy_empty_conservation.parquet"
    _write_parquet(
        conservation,
        ", ".join(f"{field} VARCHAR" for field in POCKET_CONSERVATION_FIELDS),
        [],
    )

    query = _final_query(
        config=load_config(synthetic_config),
        prestructure=prestructure,
        conservation=conservation,
        alignment_summary=None,
    )
    connection = duckdb.connect(":memory:")
    try:
        row = connection.execute(
            "SELECT conservation_status, structural_species_fraction, "
            "three_dimensional_alignment_status, grant_aligned_final_pass "
            f"FROM ({query}) AS final"
        ).fetchone()
    finally:
        connection.close()

    assert row == ("NO_STRUCTURAL_EVIDENCE", 0.0, "NOT_ASSESSED", False)


def test_empty_conservation_output_has_typed_parquet_schema(tmp_path: Path) -> None:
    """New empty Stage 09 outputs must retain their scientific column types."""
    parquet_path = tmp_path / "pocket_conservation_summary.parquet"
    write_records(
        tsv_path=tmp_path / "pocket_conservation_summary.tsv",
        parquet_path=parquet_path,
        fieldnames=POCKET_CONSERVATION_FIELDS,
        records=[],
        column_types=POCKET_CONSERVATION_COLUMN_TYPES,
    )
    connection = duckdb.connect(":memory:")
    try:
        schema = {
            str(row[0]): str(row[1])
            for row in connection.execute(
                "DESCRIBE SELECT * FROM read_parquet(?)",
                [str(parquet_path)],
            ).fetchall()
        }
    finally:
        connection.close()

    assert schema["structured_species_count"] == "BIGINT"
    assert schema["conserved_component_fraction"] == "DOUBLE"
    assert schema["all_assessed_members_pass_mapping"] == "BOOLEAN"


def test_downloaded_evidence_to_app_ready_release(
    synthetic_config: Path,
    package_root: Path,
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    """All scientific adapters share one flexible two-species schema."""
    base = load_config(synthetic_config)
    candidate_root = base.run_root / "03_candidate_evidence" / "candidate_evidence"
    candidate_schema = (
        "representative_id VARCHAR, matched_seed_ids_calculated VARCHAR, "
        "matched_seed_id_count INTEGER, reviewed_seed_count INTEGER, "
        "ubiquitin_go_positive_seed_count INTEGER, "
        "seed_with_exclusion_go_term_count INTEGER, strict_member_count INTEGER, "
        "strict_named_species_count INTEGER, strict_named_proteome_count INTEGER, "
        "strict_onekp_species_count INTEGER, seed_categories VARCHAR, "
        "seed_protein_names VARCHAR"
    )
    _write_parquet(
        candidate_root / "e3_cluster_candidate_evidence.parquet",
        candidate_schema,
        [("cluster_1", "Q9SA03;Q00002", 2, 2, 2, 0, 10, 2, 2, 0, "F-box", "fixture")],
    )

    orthology = base.run_root / "05_orthology" / "orthology" / "tables"
    _write_parquet(
        orthology / "candidate_membership_mapping.parquet",
        (
            "cluster_id VARCHAR, candidate_accession VARCHAR, record_type VARCHAR, "
            "group_id VARCHAR, species VARCHAR, mapping_status VARCHAR, "
            "ambiguity_status VARCHAR"
        ),
        [
            (
                "cluster_1",
                "Q9SA03",
                "ORTHOGROUP",
                "OG0001",
                "Arabidopsis_thaliana",
                "MATCHED",
                "UNAMBIGUOUS",
            ),
            (
                "cluster_1",
                "Q00002",
                "ORTHOGROUP",
                "OG0001",
                "Oryza_sativa",
                "MATCHED",
                "UNAMBIGUOUS",
            ),
            (
                "cluster_1",
                "Q9SA03",
                "HIERARCHICAL_ORTHOGROUP",
                "N0.HOG0001",
                "Arabidopsis_thaliana",
                "MATCHED",
                "UNAMBIGUOUS",
            ),
            (
                "cluster_1",
                "Q00002",
                "HIERARCHICAL_ORTHOGROUP",
                "N0.HOG0001",
                "Oryza_sativa",
                "MATCHED",
                "UNAMBIGUOUS",
            ),
        ],
    )
    membership_schema = (
        "record_type VARCHAR, group_id VARCHAR, orthogroup_id VARCHAR, "
        "gene_tree_parent_clade VARCHAR, species VARCHAR, raw_identifier VARCHAR, "
        "parsed_accession VARCHAR, parsed_entry VARCHAR, review_status VARCHAR, "
        "identifier_format VARCHAR, mapping_status VARCHAR, mapping_reason VARCHAR, "
        "source_file VARCHAR, source_row INTEGER"
    )
    _write_parquet(orthology / "orthogroup_membership.parquet", membership_schema, [])
    _write_parquet(
        orthology / "hierarchical_membership.parquet",
        membership_schema,
        [
            _membership_row(
                "Arabidopsis_thaliana", "sp|Q9SA03|FB27_ARATH", "Q9SA03", "FB27_ARATH"
            ),
            _membership_row(
                "Oryza_sativa", "tr|Q00002|ENTRY_RICE", "Q00002", "ENTRY_RICE"
            ),
        ],
    )
    _write_parquet(
        orthology / "candidate_cluster_orthology_summary.parquet",
        "cluster_id VARCHAR, mapped_candidate_count INTEGER",
        [("cluster_1", 2)],
    )
    _write_parquet(
        orthology / "candidate_group_member_sequences.parquet",
        (
            "cluster_id VARCHAR, record_type VARCHAR, group_id VARCHAR, "
            "orthogroup_id VARCHAR, species VARCHAR, internal_id VARCHAR, "
            "parsed_accession VARCHAR, sequence_length INTEGER, "
            "sequence_sha256 VARCHAR, protein_sequence VARCHAR"
        ),
        [
            (
                "cluster_1",
                "HIERARCHICAL_ORTHOGROUP",
                "N0.HOG0001",
                "OG0001",
                "Arabidopsis_thaliana",
                "0_0",
                "Q9SA03",
                8,
                hashlib.sha256(b"MACDEFGH").hexdigest(),
                "MACDEFGH",
            ),
            (
                "cluster_1",
                "HIERARCHICAL_ORTHOGROUP",
                "N0.HOG0001",
                "OG0001",
                "Oryza_sativa",
                "0_1",
                "Q00002",
                8,
                hashlib.sha256(b"MACDEFGH").hexdigest(),
                "MACDEFGH",
            ),
        ],
    )

    cache_root = tmp_path / "interpro_cache"
    cache_root.mkdir()
    for accession in ("Q9SA03", "Q00002"):
        (cache_root / f"{accession}.json").write_text(
            json.dumps(_annotation(accession)), encoding="utf-8"
        )
    domain_manifest = build_domain_cache_manifest(
        cache_root=cache_root,
        output_path=tmp_path / "domain_manifest.tsv",
    )

    sqlite_path = tmp_path / "e3.db"
    sqlite_connection = sqlite3.connect(sqlite_path)
    sqlite_connection.execute(
        "CREATE TABLE e3 (entry TEXT, gene_names TEXT, organism TEXT, entry_name TEXT)"
    )
    sqlite_connection.executemany(
        "INSERT INTO e3 VALUES (?, ?, ?, ?)",
        [
            ("Q9SA03", "AT1G31090", "Arabidopsis thaliana", "FB27_ARATH"),
            ("Q00002", "LOC_Os01g01010", "Oryza sativa", "ENTRY_RICE"),
        ],
    )
    sqlite_connection.commit()
    sqlite_connection.close()

    expression_root = tmp_path / "expression"
    expression_schema = (
        "experiment_accession VARCHAR, species_column VARCHAR, gene_id VARCHAR, "
        "gene_name VARCHAR, sample_or_condition VARCHAR, expression_value DOUBLE, "
        "expression_minimum DOUBLE, expression_lower_quartile DOUBLE, "
        "expression_median DOUBLE, expression_upper_quartile DOUBLE, "
        "expression_maximum DOUBLE, expression_value_statistic VARCHAR, "
        "expression_summary_type VARCHAR, expression_unit VARCHAR, source_file VARCHAR, "
        "source_file_sha256 VARCHAR"
    )
    for species, gene in (
        ("Arabidopsis_thaliana", "AT1G31090"),
        ("Oryza_sativa", "LOC_Os01g01010"),
    ):
        _write_parquet(
            expression_root
            / "atlas_expression_long"
            / f"species_column={species}"
            / "part.parquet",
            expression_schema,
            [
                (
                    "E-MTAB-1",
                    species,
                    gene,
                    gene,
                    "leaf",
                    5.0,
                    5.0,
                    5.0,
                    5.0,
                    5.0,
                    5.0,
                    "median",
                    "atlas_five_number_summary",
                    "TPM",
                    "fixture",
                    "a" * 64,
                )
            ],
        )
    expression_manifest = build_expression_manifest(
        expression_root=expression_root,
        output_path=tmp_path / "expression_manifest.tsv",
    )

    ligand_root = tmp_path / "ligandability" / "tables" / "parquet"
    _write_parquet(
        ligand_root / "joined_pockets.parquet",
        (
            "accession VARCHAR, pocket_number INTEGER, druggability_score DOUBLE, "
            "p2rank_score DOUBLE, p2rank_probability DOUBLE, p2rank_match_status VARCHAR"
        ),
        [("Q9SA03", 1, 0.9, 0.8, 0.8, "MATCHED"), ("Q00002", 1, 0.85, 0.8, 0.8, "MATCHED")],
    )
    _write_parquet(
        ligand_root / "pocket_quality.parquet",
        (
            "accession VARCHAR, pocket_number INTEGER, mapping_fraction DOUBLE, "
            "conservative_fraction_plddt_ge_70 DOUBLE, mapped_mean_plddt DOUBLE"
        ),
        [("Q9SA03", 1, 1.0, 0.9, 90.0), ("Q00002", 1, 1.0, 0.85, 88.0)],
    )
    _write_parquet(
        ligand_root / "pocket_residue_mappings.parquet",
        (
            "accession VARCHAR, pocket_number INTEGER, mapping_status VARCHAR, "
            "model_label_chain VARCHAR, model_label_seq_id INTEGER, "
            "model_auth_chain VARCHAR, model_auth_seq_id INTEGER, "
            "model_insertion_code VARCHAR, model_residue_name VARCHAR, "
            "model_plddt DOUBLE"
        ),
        [
            ("Q9SA03", 1, "MAPPED", "A", 2, "A", 2, "", "ALA", 90.0),
            ("Q9SA03", 1, "MAPPED", "A", 3, "A", 3, "", "CYS", 90.0),
            ("Q00002", 1, "MAPPED", "A", 2, "A", 2, "", "ALA", 88.0),
            ("Q00002", 1, "MAPPED", "A", 3, "A", 3, "", "CYS", 88.0),
        ],
    )
    _write_parquet(
        ligand_root / "model_quality.parquet",
        "accession VARCHAR, mean_plddt DOUBLE",
        [("Q9SA03", 90.0), ("Q00002", 88.0)],
    )
    ligandability_manifest = build_ligandability_manifest(
        roots=[ligand_root.parents[1]],
        output_path=tmp_path / "ligandability_manifest.tsv",
    )

    working = base.run_root / "04_orthofinder" / "Results" / "WorkingDirectory"
    working.mkdir(parents=True)
    (working / "SequenceIDs.txt").write_text(
        "0_0: sp|Q9SA03|FB27_ARATH\n0_1: tr|Q00002|ENTRY_RICE\n",
        encoding="utf-8",
    )
    (working / "Species0.fa").write_text(
        ">0_0\nMACDEFGH\n>0_1\nMACDEFGH\n", encoding="utf-8"
    )

    target_species = ("Arabidopsis_thaliana", "Oryza_sativa")
    prioritisation = replace(
        base.analysis.prioritisation,
        target_species=target_species,
        mandatory_species=target_species,
        minimum_target_species_fraction=1.0,
        minimum_expression_species_fraction=1.0,
        minimum_domain_species_fraction=1.0,
        minimum_structural_species_fraction=1.0,
        structure_group_limit=1,
        final_candidate_limit=1,
    )
    domains = replace(
        base.analysis.domains,
        mode="downloaded_manifest",
        allow_network=False,
    )
    config = replace(
        base,
        mode="production",
        resources=replace(
            base.resources,
            inherited_sqlite=sqlite_path,
            expression_manifest=expression_manifest,
            ligandability_manifest=ligandability_manifest,
            domain_annotation_manifest=domain_manifest,
            e3_domain_catalogue=package_root / "data" / "e3_domain_catalogue.tsv",
        ),
        analysis=replace(
            base.analysis,
            domains=domains,
            prioritisation=prioritisation,
        ),
    )

    run_domain_stage(config=config, stage_root=config.run_root / "06_domains")
    run_expression_stage(config=config, stage_root=config.run_root / "07_expression")
    run_prestructure_stage(config=config, stage_root=config.run_root / "08_shortlist_gate")

    def copy_alignment(**kwargs: Any) -> None:
        """Use the identical fixture sequences as their own deterministic alignment."""
        shutil.copyfile(kwargs["input_fasta"], kwargs["output_fasta"])
        kwargs["log_path"].write_text("fixture alignment\n", encoding="utf-8")

    monkeypatch.setattr("e3workflow.ligandability._run_mafft", copy_alignment)
    run_ligandability_stage(
        config=config, stage_root=config.run_root / "09_ligandability"
    )
    structural_tables = (
        config.run_root
        / "09b_structural_alignment"
        / "structural_alignment"
        / "tables"
    )
    _write_parquet(
        structural_tables / "structural_alignments.parquet",
        "cluster_id VARCHAR, alignment_tool VARCHAR",
        [("cluster_1", "US-align"), ("cluster_1", "TM-align")],
    )
    _write_parquet(
        structural_tables / "pocket_comparisons.parquet",
        "cluster_id VARCHAR, alignment_tool VARCHAR",
        [("cluster_1", "US-align"), ("cluster_1", "TM-align")],
    )
    _write_parquet(
        structural_tables / "pocket_residue_matches.parquet",
        "cluster_id VARCHAR, alignment_tool VARCHAR",
        [("cluster_1", "US-align"), ("cluster_1", "TM-align")],
    )
    _write_parquet(
        structural_tables / "structural_alignment_summary.parquet",
        (
            "cluster_id VARCHAR, primary_group_type VARCHAR, primary_group_id VARCHAR, "
            "three_dimensional_pocket_score DOUBLE, "
            "position_alignment_status VARCHAR, alignment_status VARCHAR, "
            "mean_minimum_tm_score DOUBLE, "
            "mean_pocket_overlap_fraction DOUBLE, "
            "median_centroid_distance_angstrom DOUBLE, "
            "mean_structural_residue_match_fraction DOUBLE, "
            "mean_structural_residue_identity_fraction DOUBLE, "
            "mean_structural_chemical_group_conservation DOUBLE"
        ),
        [
            (
                "cluster_1",
                "HIERARCHICAL_ORTHOGROUP",
                "N0.HOG0001",
                0.9,
                "SAME_3D_POCKET_POSITION_SUPPORTED",
                "CONSERVED_3D_POCKET_SUPPORTED",
                0.9,
                0.9,
                1.0,
                0.9,
                0.8,
                0.9,
            )
        ],
    )
    config = replace(
        config,
        stages=tuple(
            replace(
                stage,
                enabled=True,
                required=False,
                evidence_mode="generate",
            )
            if stage.name == "09b_structural_alignment"
            else stage
            for stage in config.stages
        ),
    )
    run_integrated_stage(
        config=config, stage_root=config.run_root / "10_integrated_resource"
    )
    run_app_ready_stage(config=config, stage_root=config.run_root / "11_app_ready")

    database = config.run_root / "10_integrated_resource/duckdb/e3_integrated_resource.duckdb"
    connection = duckdb.connect(str(database), read_only=True)
    try:
        result = connection.execute(
            "SELECT recommendation_status, grant_aligned_final_pass, "
            "orthofinder_orthogroup_ids, orthofinder_hierarchical_group_ids, "
            "three_dimensional_position_status "
            "FROM final_candidate_prioritisation"
        ).fetchone()
        tables = {
            row[0]
            for row in connection.execute("SHOW TABLES").fetchall()
        }
    finally:
        connection.close()
    assert result == (
        "PRIORITY_RECOMMENDATION",
        True,
        "OG0001",
        "N0.HOG0001",
        "SAME_3D_POCKET_POSITION_SUPPORTED",
    )
    assert {
        "domain_summary",
        "candidate_expression_summary",
        "selected_pockets",
        "pocket_sequence_coordinates",
        "candidate_group_member_sequences",
        "candidate_master_results",
        "structural_alignment_summary",
        "structural_pocket_residue_matches",
    }.issubset(tables)
    master = (
        config.run_root
        / "10_integrated_resource/tables/e3_candidate_master_results.parquet"
    )
    assert master.is_file()
    master_row = duckdb.connect(":memory:").execute(
        "SELECT cluster_id, discovery_matched_seed_ids_calculated, "
        "orthofinder_group_member_count, selected_pocket_count "
        "FROM read_parquet(?)",
        [str(master)],
    ).fetchone()
    assert master_row == ("cluster_1", "Q9SA03;Q00002", 2, 2)
    assert (
        config.run_root
        / "10_integrated_resource/reports/final_computational_prioritisation.html"
    ).is_file()
    final_root = config.run_root / "10_integrated_resource/final_results"
    assert (final_root / "top_20_computational_review_shortlist.parquet").is_file()
    assert (final_root / "top_computational_review_shortlist.parquet").is_file()
    assert (final_root / "gate_sensitivity_summary.parquet").is_file()
    assert (final_root / "grant_aligned_predicted_candidates.parquet").is_file()
    workbook = load_workbook(
        final_root / "final_candidate_recommendations.xlsx",
        read_only=False,
        data_only=False,
    )
    try:
        assert "Top_Review" in workbook.sheetnames
        assert "Gate_Sensitivity" in workbook.sheetnames
        assert "Evolutionary_Groups" in workbook.sheetnames
        assert all(
            worksheet.freeze_panes == "A2"
            for worksheet in workbook.worksheets
        )
    finally:
        workbook.close()
    assert (config.run_root / "11_app_ready/app_release_manifest.json").is_file()
    assert (
        config.run_root / "11_app_ready/config/python_app_master_parquet.env"
    ).is_file()
    assert "E3_MAX_TABLE_ROWS=10000" in (
        config.run_root / "11_app_ready/config/python_app.env"
    ).read_text(encoding="utf-8")
    assert "E3_FINAL_RESULTS_DIR=" in (
        config.run_root / "11_app_ready/config/python_app.env"
    ).read_text(encoding="utf-8")
