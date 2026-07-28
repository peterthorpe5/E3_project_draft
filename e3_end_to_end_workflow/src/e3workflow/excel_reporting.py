"""Well-formatted Excel reporting for final E3 decision outputs."""

from __future__ import annotations

import math
from datetime import date, datetime
from pathlib import Path
from typing import Any, Mapping

import duckdb
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from e3workflow.config import WorkflowConfig
from e3workflow.errors import StageError
from e3workflow.io_utils import utc_now
from e3workflow.tabular import quote_identifier

HEADER_FILL = PatternFill(fill_type="solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
PASS_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")

WORKBOOK_RELATIONS = (
    (
        "Top_20_Review",
        "top_20_computational_review_shortlist",
        "Boss-facing top-20 computational review set.",
    ),
    (
        "Predicted_Candidates",
        "grant_aligned_predicted_candidates",
        "Groups satisfying every enabled grant-aligned evidence gate.",
    ),
    (
        "Evolutionary_Groups",
        "final_evolutionary_candidate_prioritisation",
        "One row per distinct evolutionary candidate group.",
    ),
    (
        "Cluster_Contributors",
        "final_evolutionary_group_cluster_contributors",
        "DeepClust contributors retained separately from evolutionary groups.",
    ),
    (
        "Exclusion_Audit",
        "final_candidate_exclusion_audit",
        "Explicit evidence gaps and exclusion reasons.",
    ),
    (
        "Structural_Summary",
        "structural_alignment_summary",
        "Group-level US-align and TM-align evidence.",
    ),
    (
        "Representative_Audit",
        "structural_representative_selection_audit",
        "Primary and alternative accession selection metadata.",
    ),
    (
        "Pocket_Evidence",
        "selected_pockets",
        "Selected targetable-pocket evidence.",
    ),
    (
        "All_Cluster_Results",
        "candidate_master_results",
        "Complete cluster-level integrated metadata.",
    ),
    (
        "Relation_Catalogue",
        "resource_relation_catalog",
        "Normalised DuckDB relation catalogue and provenance.",
    ),
)


def _normalise_cell(value: Any) -> Any:
    """Return one Excel-safe scalar without changing numeric types."""
    if value is None:
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, (str, int, float, bool, date, datetime)):
        return value
    if isinstance(value, (list, tuple, set)):
        return ";".join(str(item) for item in value)
    return str(value)


def _column_number_format(name: str) -> str:
    """Return a readable format based on a stable scientific field name."""
    lower = name.lower()
    if lower.endswith("_fraction") or "fraction_" in lower:
        return "0.0%"
    if lower.endswith("_count") or lower.endswith("_rank") or lower.endswith("_length"):
        return "#,##0"
    if (
        lower.endswith("_score")
        or lower.endswith("_angstrom")
        or lower.endswith("_distance")
        or lower.endswith("_ratio")
    ):
        return "0.000"
    return "General"


def _table_name(sheet_name: str) -> str:
    """Return a unique Excel-safe table name."""
    return "E3_" + "".join(
        character if character.isalnum() else "_"
        for character in sheet_name
    )


def _relation_names(connection: duckdb.DuckDBPyConnection) -> set[str]:
    """Return available main-schema tables and views."""
    return {
        str(row[0])
        for row in connection.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema = 'main'"
        ).fetchall()
    }


def _write_relation_sheet(
    *,
    workbook: Workbook,
    connection: duckdb.DuckDBPyConnection,
    sheet_name: str,
    relation: str,
) -> int:
    """Write one DuckDB relation as a filtered, frozen Excel table."""
    worksheet = workbook.create_sheet(sheet_name)
    cursor = connection.execute(
        f"SELECT * FROM {quote_identifier(relation)}"
    )
    columns = [str(item[0]) for item in cursor.description]
    if not columns:
        raise StageError(f"Relation contains no columns: {relation}")
    worksheet.append(columns)
    widths = [max(10, min(60, len(column) + 2)) for column in columns]
    row_count = 0
    while True:
        batch = cursor.fetchmany(1_000)
        if not batch:
            break
        for row in batch:
            values = [_normalise_cell(value) for value in row]
            worksheet.append(values)
            row_count += 1
            if row_count <= 500:
                for index, value in enumerate(values):
                    if value is None:
                        continue
                    widths[index] = max(
                        widths[index],
                        min(60, len(str(value)) + 2),
                    )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 28
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    for index, column in enumerate(columns, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = widths[index - 1]
        number_format = _column_number_format(column)
        if number_format != "General":
            for cell in worksheet.iter_cols(
                min_col=index,
                max_col=index,
                min_row=2,
                max_row=max(2, row_count + 1),
            ):
                cell[0].number_format = number_format
        if any(
            token in column.lower()
            for token in (
                "reason",
                "interpretation",
                "accessions",
                "species_present",
                "species_missing",
                "metadata",
            )
        ):
            worksheet.column_dimensions[get_column_letter(index)].width = min(
                60,
                max(35, widths[index - 1]),
            )
    if row_count > 0:
        table = Table(
            displayName=_table_name(sheet_name),
            ref=f"A1:{get_column_letter(len(columns))}{row_count + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    return row_count


def _write_readme_sheet(
    *,
    workbook: Workbook,
    config: WorkflowConfig,
    relation_counts: Mapping[str, int],
) -> None:
    """Write the workbook interpretation and navigation sheet."""
    worksheet = workbook.create_sheet("README", 0)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    rows = [
        ("ARIA plant E3 structural-completion results", ""),
        ("Run", config.run_name),
        ("Generated", utc_now()),
        ("Scoring profile", config.analysis.prioritisation.profile_name),
        ("Top review limit", config.analysis.prioritisation.final_candidate_limit),
        ("Structural group limit", config.analysis.prioritisation.structure_group_limit),
        (
            "Official ranking treatment",
            (
                "The existing evidence-weighted rank is retained. Three-dimensional "
                "alignment is an explicit eligibility/support gate and is not silently "
                "reweighted unless its thresholds are reviewed."
            ),
        ),
        (
            "Interpretation",
            (
                "GRANT_ALIGNED_PREDICTED_CANDIDATE means that every configured "
                "computational gate passed. It does not prove E3 activity, ligand "
                "binding, selectivity or target degradation."
            ),
        ),
        (
            "Evolutionary group identity",
            (
                "The principal recommendation tables contain one row per distinct "
                "primary OrthoFinder group. Contributing DeepClust clusters remain "
                "separate in Cluster_Contributors."
            ),
        ),
        ("Sheet", "Rows"),
    ]
    rows.extend((name, count) for name, count in relation_counts.items())
    for row in rows:
        worksheet.append(row)
    worksheet["A1"].fill = HEADER_FILL
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    worksheet["B1"].fill = HEADER_FILL
    for cell in worksheet[10]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)
    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 92
    for row in worksheet.iter_rows(min_row=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[7].height = 48
    worksheet.row_dimensions[8].height = 60
    worksheet.row_dimensions[9].height = 54


def _write_settings_sheet(
    *,
    workbook: Workbook,
    config: WorkflowConfig,
) -> None:
    """Write auditable thresholds and decision settings."""
    structural = config.analysis.structural_alignment
    ligandability = config.analysis.ligandability
    prioritisation = config.analysis.prioritisation
    records = [
        ("Target species", ";".join(prioritisation.target_species)),
        ("Mandatory species", ";".join(prioritisation.mandatory_species)),
        ("Minimum target-species fraction", prioritisation.minimum_target_species_fraction),
        (
            "Minimum expression-species fraction",
            prioritisation.minimum_expression_species_fraction,
        ),
        ("Minimum domain-species fraction", prioritisation.minimum_domain_species_fraction),
        (
            "Minimum structural-species fraction",
            prioritisation.minimum_structural_species_fraction,
        ),
        ("Minimum druggability score", ligandability.minimum_druggability_score),
        ("Minimum mapping fraction", ligandability.minimum_mapping_fraction),
        ("Minimum pocket-pLDDT fraction", ligandability.minimum_pocket_plddt_fraction),
        ("Minimum aligned-region overlap", ligandability.minimum_region_overlap),
        ("Minimum global TM-score", structural.minimum_global_tm_score),
        (
            "Maximum centroid distance (Å)",
            structural.maximum_centroid_distance_angstrom,
        ),
        (
            "Minimum 3D pocket overlap",
            structural.minimum_pocket_overlap_fraction,
        ),
        (
            "Minimum structural residue match",
            structural.minimum_structural_residue_match_fraction,
        ),
        (
            "Minimum structural chemical-group conservation",
            structural.minimum_structural_chemical_group_conservation,
        ),
        ("Minimum group 3D support", structural.minimum_group_support_fraction),
        ("3D values reweight ranking", structural.use_for_prioritisation),
        (
            "3D support required for final recommendation",
            structural.require_for_final_recommendation,
        ),
    ]
    worksheet = workbook.create_sheet("Settings")
    worksheet.append(("Setting", "Value"))
    for record in records:
        worksheet.append(record)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    worksheet.column_dimensions["A"].width = 52
    worksheet.column_dimensions["B"].width = 90
    for cell in worksheet["B"][1:]:
        if isinstance(cell.value, float) and 0.0 <= cell.value <= 1.0:
            cell.number_format = "0.0%"
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def create_final_results_workbook(
    *,
    connection: duckdb.DuckDBPyConnection,
    config: WorkflowConfig,
    output_path: Path,
) -> Path:
    """Create and validate the final multi-sheet Excel workbook."""
    destination = Path(output_path).expanduser().resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    available = _relation_names(connection)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "ARIA E3 end-to-end workflow"
    workbook.properties.title = "ARIA plant E3 structural-completion results"
    workbook.properties.subject = "Grant-aligned computational candidate prioritisation"
    relation_counts: dict[str, int] = {}
    for sheet_name, relation, _description in WORKBOOK_RELATIONS:
        if relation not in available:
            continue
        relation_counts[sheet_name] = _write_relation_sheet(
            workbook=workbook,
            connection=connection,
            sheet_name=sheet_name,
            relation=relation,
        )
    _write_readme_sheet(
        workbook=workbook,
        config=config,
        relation_counts=relation_counts,
    )
    _write_settings_sheet(workbook=workbook, config=config)
    temporary = destination.with_name(f".{destination.name}.partial")
    temporary.unlink(missing_ok=True)
    workbook.save(temporary)
    temporary.replace(destination)
    try:
        verified = load_workbook(destination, read_only=False, data_only=False)
        if "README" not in verified.sheetnames or "Top_20_Review" not in verified.sheetnames:
            raise StageError("Final Excel workbook lacks required sheets")
        for worksheet in verified.worksheets:
            if worksheet.freeze_panes != "A2":
                raise StageError(
                    f"Final Excel sheet does not freeze its top row: {worksheet.title}"
                )
        verified.close()
    except (OSError, ValueError) as exc:
        raise StageError(f"Could not validate final Excel workbook: {exc}") from exc
    return destination
